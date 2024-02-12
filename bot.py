import time

import telebot
from telebot import types
import datetime
from datetime import datetime, timedelta
import schedule
from openpyxl import load_workbook
import os
from decouple import config

TOKEN = config('TELEGRAM_BOT_TOKEN')
bot = telebot.TeleBot(TOKEN)
filename = "attendance.xlsx"
bonus_plan = "план продаж.xlsx"
individual_bonus = 0
group_chat_id = 0

import requests
from requests.adapters import HTTPAdapter
from urllib3.util.retry import Retry

def send_telegram_request(url, max_retries=3, backoff_factor=1):
    session = requests.Session()
    retries = Retry(total=max_retries,
                    backoff_factor=backoff_factor,
                    status_forcelist=[500, 502, 503, 504],
                    allowed_methods=frozenset(['GET', 'POST']))
    adapter = HTTPAdapter(max_retries=retries)
    session.mount('http://', adapter)
    session.mount('https://', adapter)

    try:
        response = session.get(url)  # Используйте session.post(url, data) для POST-запроса
        response.raise_for_status()  # Вызовет исключение для HTTP-кодов ошибок
        return response.json()  # Возвращает JSON ответа
    except requests.exceptions.HTTPError as http_err:
        print(f'HTTP error occurred: {http_err}')  # Например, "404 Not Found"
    except requests.exceptions.ConnectionError as conn_err:
        print(f'Connection error occurred: {conn_err}')
    except requests.exceptions.Timeout as timeout_err:
        print(f'Timeout error occurred: {timeout_err}')
    except requests.exceptions.RequestException as req_err:
        print(f'Unknown error occurred: {req_err}')


@bot.message_handler(commands=['start'], func=lambda message: message.chat.type == 'private')
def start_message(message):
    chat_id = message.chat.id
    bot.send_message(chat_id,
                     "Привет, я бот который будет считать премию твоих сотрудников). Для начала добавь меня в группу, где сотрудники будут писать дневную выручку")


@bot.message_handler(func=lambda message: message.text.lower() == 'я на смене')
def on_arrival(message):
    global time_str
    global group_chat_id
    group_chat_id = message.chat.id
    message_datetime_utc = datetime.utcfromtimestamp(message.date)
    message_datetime_moscow = message_datetime_utc + timedelta(hours=3)
    time_str = message_datetime_moscow.strftime('%H:%M')
    keyboard = types.InlineKeyboardMarkup(row_width=2)

    confirm_button = types.InlineKeyboardButton("Подтвердить ✅", callback_data='confirm')
    keyboard.add(confirm_button)

    cancel_button = types.InlineKeyboardButton("Отменить ❌", callback_data='cancel')
    keyboard.add(cancel_button)

    bot.send_message(group_chat_id, "Отлично, рад видеть вас на месте!", reply_markup=keyboard)


bonus_data = None


@bot.callback_query_handler(func=lambda call: True)
def handle_callback_query(call):
    global time_str
    group_chat_id = call.message.chat.id
    message_id = call.message.message_id
    if call.data == 'confirm':
        bot.send_message(call.message.chat.id,
                         f" Привет!  @{call.from_user.username}! Хорошего рабочего дня и отличного настроения :)")

        write_arrival_and_bonus_to_excel(filename, call.from_user.username, time_str, bonus=0)

    bot.delete_message(chat_id=group_chat_id, message_id=message_id)


@bot.message_handler(content_types=['photo'])
def handle_photo_report(message):


    User = message.from_user.username
    if message.caption and message.caption.startswith('#отчёт') or message.caption.startswith('#отчет'):
        try:
            individual_bonus = 0
            group_chat_id = message.chat.id
            parts = message.caption.split()
            if len(parts) != 3:
                raise ValueError("Неправильный формат команды")

            _, date_str, revenue_str = parts

            # Преобразуем строку с датой в объект datetime и добавляем текущий год
            current_year = datetime.now().year
            report_date = datetime.strptime(date_str, '%d.%m').replace(year=current_year).date()
            print(report_date)
            today = datetime.today().date()
            bonus_value = get_bonus_from_excel(float(revenue_str), report_date)
            num_employees = get_num_employees(report_date)
            # Проверяем, соответствует ли дата сообщения текущему дню
            if report_date <= today:
                print("bonus_value:", bonus_value)
                print("num:",num_employees)
                individual_bonus = int(bonus_value / num_employees) if num_employees > 0 else 0
                print(individual_bonus)
                # Обновляем четвертую колонку для всех соответствующих строк

                write_arrival_and_bonus_to_excel(filename, username=User, arrival_time=None,
                                                 bonus=individual_bonus,date=report_date)
                print("Сохранено")

            report_daily_bonus(group_chat_id, individual_bonus)
        except ValueError as e:
            bot.send_message(message.chat.id, "Неправильный формат отчета. Используйте: #отчёт ДД.ММ СУММА")
    else:
        pass


def is_weekday(date):
    log_time= datetime.now()
    log_time = log_time.strftime("%d/%m/%Y %H:%M:%S")
    print(log_time, ":", date)

    # Получаем номер текущего дня недели (0 - понедельник, 1 - вторник, и так далее)
    day_of_week = date.weekday()

    # Проверяем, является ли текущий день будним (понедельник - пятница)
    if 0 <= day_of_week <= 3:
        return 1
    else:
        return 0

def get_num_employees(date):
        date = date.day
        log_time= datetime.now()
        log_time = log_time.strftime("%d/%m/%Y %H:%M:%S")
        print(log_time, ":", date)


        workbook = load_workbook(filename)
        sheet = workbook.active
        # Находим столбец, который соответствует сегодняшнему дню
        today_column = None
        for col in range(1, sheet.max_column + 1):
            if sheet.cell(row=1, column=col).value == date:
                today_column = col
                break

        # Если столбец с сегодняшней датой найден, считаем количество сотрудников на смене
        print(today_column)
        num_employees = 0
        if today_column:
            for row in range(3, sheet.max_row + 1, 2):  # Пропускаем одну строку между записями сотрудников
                if sheet.cell(row=row, column=today_column).value is not None:
                    num_employees += 1
                    print(num_employees)
        return num_employees



def write_arrival_and_bonus_to_excel(filename, username, arrival_time, bonus=None, date=None):
    if not os.path.exists(filename):
        print(f"Файл {filename} не найден.")
        return
    workbook = load_workbook(filename)
    sheet = workbook.active
    if date is None:
        today = datetime.today()
        date = today.day
    else:
        date = date.day
    log_time= datetime.now()
    log_time = log_time.strftime("%d/%m/%Y %H:%M:%S")
    print(log_time, ":", date)


    user_row = None
    for row in range(3, sheet.max_row + 1):
        if sheet.cell(row=row, column=1).value == username:
            user_row = row
            break

    if user_row is None:
        print(f"Пользователь {username} не найден.")
        return

    day_col = None
    for col in range(2, sheet.max_column + 1):
        if sheet.cell(row=1, column=col).value == date:
            day_col = col
            break

    if day_col is None:
        print(f"Для текущего дня {date} не найдена колонка.")
        return

    # Запись времени прибытия, если указано
    if arrival_time:
        sheet.cell(row=user_row, column=day_col, value=arrival_time)
        sheet.cell(row=user_row + 1, column=day_col, value=bonus)
    # Запись премии на строку ниже, если указана
    if bonus is not None:
        for row in range(3, sheet.max_row + 1, 2):  # Пропускаем одну строку между записями сотрудников
            if sheet.cell(row=row, column=day_col).value is not None:
                sheet.cell(row=row+1, column=day_col, value=bonus)

    workbook.save(filename)



def get_bonus_from_excel(revenue, date):
    # Определяем, является ли сегодня будним или выходным
    weekday = is_weekday(date)  # 1 для буднего, 0 для выходного

    revenue_col_idx = 1 if weekday else 4
    bonus_col_idx = 2 if weekday else 5

    workbook = load_workbook(bonus_plan)
    sheet = workbook.active

    rows = list(sheet.iter_rows(min_row=4, max_col=bonus_col_idx, values_only=True))
    for i in range(len(rows) - 1):  # Итерируем до предпоследней строки
        current_row = rows[i]
        next_row = rows[i + 1]

        threshold = current_row[revenue_col_idx - 1]
        next_threshold = next_row[revenue_col_idx - 1]
        bonus = current_row[bonus_col_idx - 1]

        # Проверяем, попадает ли revenue в диапазон между текущим и следующим порогом
        # Для последней строки, next_threshold может быть не определён, поэтому используем float('inf')
        if threshold is not None and threshold <= revenue < (
        next_threshold if next_threshold is not None else float('inf')):
            print("bonus", bonus)
            return bonus

    # Обрабатываем случай для последней строки отдельно
    if rows and revenue >= rows[-1][revenue_col_idx - 1]:
        return rows[-1][bonus_col_idx - 1]

    return 0  # Возвращает 0, если не найдено подходящее значение


@bot.message_handler(commands=['bonus'], func=lambda message: message.chat.type == 'private')
def handle_bonus_command(message):
    # Отправляем запрос на отправку файла 'план продаж.xlsx'
    bot.send_message(message.chat.id, "Пожалуйста, отправьте файл 'план продаж.xlsx' с данными о выручке и премиях.")


@bot.message_handler(content_types=['document'], func=lambda message: message.chat.type == 'private')
def handle_document(message):
    if message.document.file_name == 'план продаж.xlsx':
        try:
            # Открываем файл Excel
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            with open('план продаж.xlsx', 'wb') as new_file:
                new_file.write(downloaded_file)

        except Exception as e:
            bot.send_message(message.chat.id, "Произошла ошибка при обработке файла. Пожалуйста, попробуйте еще раз.")
    else:
        bot.send_message(message.chat.id, "Пожалуйста, отправьте файл 'план продаж.xlsx'.")
    # load_bonus_data()


def report_daily_bonus(group_chat_id, individual_bonus):
    message = f"Премия каждого сотрудника: {individual_bonus}"

    bot.send_message(group_chat_id, message)


@bot.message_handler(commands=['clear'])
def clear_excel_table(message):
    try:
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Очистка таблицы (удаление всех значений)
        for row in sheet.iter_rows(min_row=3, max_row=sheet.max_row):
            for cell in row:
                cell.value = None

        workbook.save(filename)
        bot.reply_to(message, "Таблица успешно очищена!")
    except Exception as e:
        bot.reply_to(message, f"Произошла ошибка: {str(e)}")


@bot.message_handler(commands=['download'])
def download_excel_file(message):
    try:
        with open(filename, 'rb') as file:
            bot.send_document(message.chat.id, file)
    except Exception as e:
        bot.reply_to(message, f"Произошла ошибка: {str(e)}")


@bot.message_handler(commands=['add'])
def handle_add(message):
    msg = bot.send_message(message.chat.id, "Введите теги сотрудников, разделенные пробелом")
    bot.register_next_step_handler(msg, process_user_tags)

def process_user_tags(message):
    # Разделяем полученный текст на теги по пробелам
    user_tags = message.text.split()

    # Удаляем символ "@" из каждого тега сотрудника и обрабатываем каждый тег
    for tag in user_tags:
        username = tag.replace("@", "")  # Удаление "@" из тега
        workbook = load_workbook(filename)
        sheet = workbook.active

        # Используем функцию find_first_empty_row для определения первой пустой строки
        max_row = find_first_empty_row(sheet)

        # Добавляем имя пользователя и слово "Премия" на последующих строках
        sheet[f'A{max_row}'] = username
        sheet[f'A{max_row + 1}'] = "Премия"

        workbook.save(filename)

        bot.send_message(message.chat.id, f"Тег сотрудника {username} обработан.")

    # Можно добавить сообщение о завершении обработки всех тегов
    bot.send_message(message.chat.id, "Все теги сотрудников были обработаны.")

def find_first_empty_row(sheet):
    for row in range(1, sheet.max_row + 1):
        if not any(cell.value for cell in sheet[row]):
            return row
    return sheet.max_row + 1

if __name__ == "__main__":
    connected = False
while not connected:
    try:
        bot.polling()
        connected = True
    except Exception as e:
        print(f"Error: {e}")
        print("Trying to reconnect in 10 seconds...")
        time.sleep(10)

while True:
    schedule.run_pending()
    time.sleep(1)