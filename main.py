import time

import telebot
from telebot import types
import datetime
from datetime import datetime, timedelta
import schedule
from openpyxl import load_workbook
from openpyxl import Workbook
import os
from decouple import config

TOKEN = config('TELEGRAM_BOT_TOKEN')
bot = telebot.TeleBot(TOKEN)
filename = "attendance.xlsx"
bonus_plan = "план продаж.xlsx"
num_employees = 0
individual_bonus = 0
group_chat_id = 0


@bot.message_handler(commands=['start'], func=lambda message: message.chat.type == 'private')
def start_message(message):
    chat_id = message.chat.id
    bot.send_message(chat_id,
                     "Привет, я бот который будет считать премию твоих сотрудников). Для начала добавь меня в группу, где сотрудники будут писать дневную выручку")


@bot.message_handler(func=lambda message: message.text.lower() == 'я на смене')
def on_arrival(message):
    global time_str
    global group_chat_id
    group_chat_id = message.chat.id
    message_datetime_utc = datetime.utcfromtimestamp(message.date)
    message_datetime_moscow = message_datetime_utc + timedelta(hours=3)
    time_str = message_datetime_moscow.strftime('%H:%M')
    keyboard = types.InlineKeyboardMarkup(row_width=2)

    confirm_button = types.InlineKeyboardButton("Подтвердить ✅", callback_data='confirm')
    keyboard.add(confirm_button)

    cancel_button = types.InlineKeyboardButton("Отменить ❌", callback_data='cancel')
    keyboard.add(cancel_button)

    bot.send_message(group_chat_id, "Отлично, рад видеть вас на месте!", reply_markup=keyboard)


bonus_data = None


def load_bonus_data():
    global bonus_data
    workbook = load_workbook(bonus_plan)
    sheet = workbook.active
    bonus_data = []

    for row in sheet.iter_rows(min_row=4, max_row=sheet.max_row, values_only=True):
        bonus_data.append(row)


@bot.callback_query_handler(func=lambda call: True)
def handle_callback_query(call):
    global time_str
    global num_employees
    group_chat_id = call.message.chat.id
    message_id = call.message.message_id
    if call.data == 'confirm':
        bot.send_message(call.message.chat.id, f"@{call.from_user.username} пришел на работу в {time_str} ✅")

        num_employees += 1
        print(num_employees)
        write_to_excel(filename, group_chat_id, call.from_user.username, time_str)

    elif call.data == 'cancel':
        bot.send_message(call.message.chat.id, "Действие отменено ❌")

    bot.delete_message(chat_id=group_chat_id, message_id=message_id)


@bot.message_handler(func=lambda message: message.text.startswith('выручка сегодня '))
def handle_revenue_message(message):
    global num_employees
    print(num_employees)
    try:
        global bonus_data
        if bonus_data is None:
            load_bonus_data()

        revenue_str = message.text.split('выручка сегодня ')[1]
        revenue = float(revenue_str)

        weekday = is_weekday()
        column = 0 if weekday else 3

        bonus_value = 0
        for i in range(len(bonus_data) - 1):  # Итерация по всем строкам, кроме последней
            current_row = bonus_data[i]
            next_row = bonus_data[i + 1]
            min_revenue = current_row[column]
            next_min_revenue = next_row[column]

            if min_revenue is not None and revenue >= min_revenue:
                if next_min_revenue is None or revenue < next_min_revenue:
                    bonus_value = current_row[column + 1]
                    break

        if bonus_value == 0 and revenue >= bonus_data[-1][column]:  # Проверка последнего порога
            bonus_value = bonus_data[-1][column + 1]

        today = datetime.today().date()
        workbook = load_workbook(filename)
        sheet = workbook.active
        global individual_bonus
        individual_bonus = bonus_value / num_employees if num_employees > 0 else 0

        for row in range(2, sheet.max_row + 1):  # Начиная со второй строки и до конца таблицы
            cell_date = sheet.cell(row=row, column=1).value
            if cell_date and cell_date.date() == today:  # Сравниваем только даты
                sheet.cell(row=row, column=4, value=individual_bonus)

        workbook.save(filename)

    except ValueError:
        bot.send_message(message.chat.id,
                         "Некорректный формат числа. Пожалуйста, используйте числа в формате 'выручка сегодня *число*'.")


def is_weekday():
    # Получаем текущую дату
    today = datetime.today()

    # Получаем номер текущего дня недели (0 - понедельник, 1 - вторник, и так далее)
    day_of_week = today.weekday()

    # Проверяем, является ли текущий день будним (понедельник - пятница)
    if 0 <= day_of_week <= 4:
        return 1
    else:
        return 0


def write_to_excel(filename, group_chat_id, username, arrival_time):
    # Проверяем, существует ли файл. Если нет, создаем его.
    if not os.path.exists(filename):
        workbook = Workbook()
        sheet = workbook.active
        sheet.append(["UserName", "Время прибытия, Премия"])
        workbook.save(filename)
    today = datetime.today().date()
    # Загружаем существующий файл
    workbook = load_workbook(filename)
    sheet = workbook.active

    # Добавляем новую строку с данными
    sheet.append([today, username, arrival_time])
    workbook.save(filename)


@bot.message_handler(commands=['bonus'], func=lambda message: message.chat.type == 'private')
def handle_bonus_command(message):
    # Отправляем запрос на отправку файла 'план продаж.xlsx'
    bot.send_message(message.chat.id, "Пожалуйста, отправьте файл 'план продаж.xlsx' с данными о выручке и премиях.")


@bot.message_handler(content_types=['document'], func=lambda message: message.chat.type == 'private')
def handle_document(message):
    if message.document.file_name == 'план продаж.xlsx':
        try:
            # Открываем файл Excel
            file_info = bot.get_file(message.document.file_id)
            downloaded_file = bot.download_file(file_info.file_path)
            with open('план продаж.xlsx', 'wb') as new_file:
                new_file.write(downloaded_file)

        except Exception as e:
            bot.send_message(message.chat.id, "Произошла ошибка при обработке файла. Пожалуйста, попробуйте еще раз.")
    else:
        bot.send_message(message.chat.id, "Пожалуйста, отправьте файл 'план продаж.xlsx'.")
    load_bonus_data()


def report_daily_bonus():
    global group_chat_id
    global individual_bonus
    message = f"Премия каждого сотрудника: {individual_bonus}"

    bot.send_message(group_chat_id, message)


schedule.every().day.at("22:00").do(report_daily_bonus)

if __name__ == "__main__":
    bot.polling()

while True:
    schedule.run_pending()
    time.sleep(1)
